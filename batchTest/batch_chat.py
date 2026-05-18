"""
Batch calls to documentApi POST /api/chat using JSON config.
Writes an English Excel workbook: questions, answers, metrics, retrieved chunks,
manual RAG-triad + type-specific Dim4 scores, and optional LLM-as-judge (0–1).

Active DB/vector environment: `activePostgresEnvironmentId` in config and/or `--environment`.
Optional `applyAppSettings` + `appSettings` for other app fields.

Optional `evaluateAnswers`: OpenAI-compatible judge; adds sheet `Eval_Charts`.

Dependencies: pip install -r requirements.txt
Run: python batch_chat.py --config batch_config.json
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from eval_judge import judge_one_http, resolve_eval_judge_config

_SCRIPT_DIR = Path(__file__).resolve().parent
# Standard-Konfiguration liegt neben diesem Skript (funktioniert unabhängig vom Arbeitsverzeichnis).
DEFAULT_CONFIG = str(_SCRIPT_DIR / "batch_config.json")
EXCEL_MAX_CELL = 32000

# Excel: core columns + manual RAG triad + Dim4 (each 0–2; sum max 8). Headers are human-readable.
RESULT_HEADERS_CORE: tuple[str, ...] = (
    "Row no.",
    "Question ID",
    "Question type",
    "Question text",
    "Ground truth (answerField)",
    "Model answer",
    "HTTP or request error",
    "Latency (ms)",
    "Prompt tokens",
    "Completion tokens",
    "Total tokens",
    "Tokens per second",
    "Number of context chunks",
    "Retrieved chunks (summary)",
    "Retrieved chunks (full text for judge)",
)
# LLM-as-judge (0–1)
RESULT_HEADERS_AUTO: tuple[str, ...] = (
    "Judge: answer relevance (0–1)",
    "Judge: context relevance (0–1)",
    "Judge: groundedness (0–1)",
    "Judge: answer correctness (0–1)",
    "Judge notes",
    "Judge error",
)
RESULT_HEADERS_EVAL: tuple[str, ...] = (
    "Human score: context relevance (0–2)",
    "Human score: groundedness (0–2)",
    "Human score: answer relevance (0–2)",
    "Human score: Dim4 criterion (hint)",
    "Human score: Dim4 (0–2)",
    "Human score: total (0–8)",
)

# Excel display: swap columns G (7) and P (16): Judge answer relevance next to model block; HTTP error after metrics.
_JUDGE_FLOAT_DISPLAY_COLS_1BASED: tuple[int, ...] = (7, 17, 18, 19)
_JUDGE_FLOAT_NUMBER_FORMAT = "0.000000"


def _swap_results_columns_g_and_p(seq: list[Any]) -> list[Any]:
    """Reorder one row/header for Results: Excel G ↔ P (0-based indices 6 ↔ 15)."""
    if len(seq) < 16:
        return list(seq)
    s = list(seq)
    return s[:6] + [s[15]] + s[7:15] + [s[6]] + s[16:]


def _dim4_kriterium_label(qtype: str) -> str:
    t = (qtype or "").strip()
    if t == "Objective":
        return "Accuracy"
    if t == "Subjective":
        return "Completeness"
    return ""


def _eval_score_column_indices_1based(manual_prefix_col_count: int) -> tuple[int, int, int, int]:
    """1-based columns: context, groundedness, answer relevance, Dim4 score (excludes criterion text and total)."""
    o = manual_prefix_col_count
    return (o + 1, o + 2, o + 3, o + 5)


def _eval_sum_formula(data_row: int, manual_prefix_col_count: int) -> str:
    c1, c2, c3, c4 = _eval_score_column_indices_1based(manual_prefix_col_count)
    refs = ",".join(f"{get_column_letter(c)}{data_row}" for c in (c1, c2, c3, c4))
    return f"=SUM({refs})"


def _parse_auto_metric_cell(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def aggregate_auto_eval_from_rows(
    rows: list[list[Any]],
    *,
    core_col_count: int,
    n_metric_cols: int = 4,
) -> dict[str, Any]:
    """
    Eine Quelle für Mittelwerte: Spalten core…core+3 = Auto-Antwortrelevanz, Kontext, Groundedness,
    Answer correctness (0–1). Zusätzlich Objective-Accuracy-Varianten für Auswertung/Konsole.
    """
    sums = [0.0] * n_metric_cols
    counts = [0] * n_metric_cols
    by_type: dict[str, list[list[float]]] = {}
    idx_correctness = 3
    obj_corr: list[float] = []

    for row in rows:
        if len(row) < core_col_count + n_metric_cols:
            continue
        typ = str(row[2] or "").strip() or "?"
        if typ not in by_type:
            by_type[typ] = [[] for _ in range(n_metric_cols)]
        for i in range(n_metric_cols):
            v = _parse_auto_metric_cell(row[core_col_count + i])
            if v is None:
                continue
            sums[i] += v
            counts[i] += 1
            by_type[typ][i].append(v)
            if i == idx_correctness and typ == "Objective":
                obj_corr.append(v)

    def _mean(vals: list[float]) -> float | None:
        return sum(vals) / len(vals) if vals else None

    return {
        "sums": sums,
        "counts": counts,
        "by_type": by_type,
        "mean_per_metric": [
            (sums[i] / counts[i]) if counts[i] else None for i in range(n_metric_cols)
        ],
        "accuracy_objective_mean": _mean(obj_corr),
        "n_objective_scored": len(obj_corr),
    }


def _print_auto_eval_accuracy_summary(agg: dict[str, Any] | None, *, had_judge: bool) -> None:
    """Stdout summary after each run (matches Excel auto-eval aggregation)."""
    print("", flush=True)
    print("— Auto evaluation (LLM judge) —", flush=True)
    if not had_judge:
        print(
            "  Accuracy / auto metrics: not computed (evaluateAnswers: false).",
            flush=True,
        )
        return
    if agg is None:
        print("  No rows to aggregate.", flush=True)
        return
    labels_en = (
        "Answer relevance",
        "Context relevance",
        "Groundedness",
        "Answer correctness (Objective + answerField only)",
    )
    means: list[float | None] = agg["mean_per_metric"]
    counts: list[int] = agg["counts"]
    for i, lab in enumerate(labels_en):
        m = means[i]
        c = counts[i]
        if m is not None:
            print(f"  Mean {lab}: {m:.4f}  (n={c})", flush=True)
        else:
            print(f"  Mean {lab}: —  (n=0)", flush=True)
    ao = agg.get("accuracy_objective_mean")
    no = int(agg.get("n_objective_scored") or 0)
    print(
        "  Accuracy (mean of those correctness values, Objective + reference only): "
        f"{ao:.4f}  (n={no})" if ao is not None else f"  Accuracy: —  (no Objective rows with reference and score, n={no})",
        flush=True,
    )


def _append_eval_diagram_sheet(
    wb: Workbook,
    *,
    aggregated: dict[str, Any],
) -> None:
    """Mittelwerte der Auto-Metriken + Balkendiagramme (Aggregation aus Zeilen, konsistent mit Konsole)."""
    sums: list[float] = aggregated["sums"]
    counts: list[int] = aggregated["counts"]
    by_type: dict[str, list[list[float]]] = aggregated["by_type"]
    if sum(counts) == 0:
        return
    labels = (
        "Answer relevance",
        "Context relevance",
        "Groundedness",
        "Answer correctness",
    )

    ws = wb.create_sheet("Eval_Charts")
    ws["A1"] = "Metric (auto, 0–1)"
    ws["B1"] = "Mean"
    for i, lab in enumerate(labels):
        ws.cell(row=2 + i, column=1, value=lab)
        mean_v = sums[i] / counts[i] if counts[i] else None
        ws.cell(row=2 + i, column=2, value=round(mean_v, 4) if mean_v is not None else "")

    if any(counts):
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Mean LLM scores (0–1)"
        chart1.y_axis.title = "Mean"
        chart1.x_axis.title = "Metric"
        chart1.height = 10
        chart1.width = 18
        dref = Reference(ws, min_col=2, min_row=1, max_row=5)
        cref = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart1.add_data(dref, titles_from_data=True)
        chart1.set_categories(cref)
        ws.add_chart(chart1, "D2")
    else:
        ws["D2"] = "No numeric auto scores (empty cells or judge errors)."

    ao = aggregated.get("accuracy_objective_mean")
    no = int(aggregated.get("n_objective_scored") or 0)
    ws["A7"] = "Accuracy (Objective + answerField, mean correctness)"
    ws["B7"] = round(ao, 4) if ao is not None else ""
    ws["C7"] = f"n={no}"
    ws["A8"] = (
        "Note: “Judge: answer correctness (0–1)” is filled only for Objective rows with a non-empty "
        "ground truth (answerField); the Answer correctness bar is the mean over those rows only."
    )
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")

    # Zweites Diagramm: gleiche Metriken nach Frage-Typ (Mittel pro Typ)
    start_row = 12
    ws.cell(row=start_row, column=1, value="Question type")
    for j, lab in enumerate(labels, start=2):
        ws.cell(row=start_row, column=j, value=lab)
    row_off = 1
    for typ in sorted(by_type.keys()):
        arrs = by_type[typ]
        ws.cell(row=start_row + row_off, column=1, value=typ)
        for j in range(4):
            vals = arrs[j]
            m = sum(vals) / len(vals) if vals else None
            ws.cell(row=start_row + row_off, column=2 + j, value=round(m, 4) if m is not None else "")
        row_off += 1
    last = start_row + row_off - 1
    if last > start_row:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Mean scores by question type (clustered)"
        chart2.grouping = "clustered"
        chart2.height = 10
        chart2.width = 22
        d2 = Reference(ws, min_col=2, min_row=start_row, max_row=last, max_col=5)
        c2 = Reference(ws, min_col=1, min_row=start_row + 1, max_row=last)
        chart2.add_data(d2, titles_from_data=True)
        chart2.set_categories(c2)
        ws.add_chart(chart2, "D18")

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 52 if col == 1 else 16


def _append_rubric_sheet(wb: Workbook, insert_at: int | None = None) -> None:
    if insert_at is not None:
        ws = wb.create_sheet("Rubric", insert_at)
    else:
        ws = wb.create_sheet("Rubric")
    ws["A1"] = "RAG triad + Dim4 (manual entry on sheet “Results”)"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    rows: list[tuple[str, ...]] = [
        ("Criterion", "Description", "0 points", "1 point", "2 points"),
        (
            "Context relevance",
            "Relevance of retrieved document segments to the question",
            "insufficient / not relevant",
            "partially relevant / somewhat fitting",
            "fully relevant",
        ),
        (
            "Groundedness",
            "How well the answer is supported by the provided context",
            "not traceable / hallucination",
            "partially derivable from context",
            "fully grounded in context",
        ),
        (
            "Answer relevance",
            "Correctness, completeness, and clarity of the answer content",
            "wrong / unclear / insufficient",
            "partially correct or incomplete",
            "fully correct and understandable",
        ),
        (
            "Dim4: Accuracy (Objective only)",
            "Factual accuracy of the answer (vs. answerField / reference where applicable)",
            "wrong / not verifiable",
            "partially correct",
            "fully correct",
        ),
        (
            "Dim4: Completeness (Subjective only)",
            "How completely the answer covers the question",
            "major aspects missing",
            "mostly covered, with gaps",
            "content complete",
        ),
        ("", "", "", "", ""),
        ("Scale", "All manual criteria (Dim1–Dim4): uniformly 0–2.", "", "", ""),
        ("Maximum", "8 points per question (sum of the four 0–2 score columns).", "", "", ""),
    ]
    for r, row in enumerate(rows, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 30 if col == 1 else 40
    ws.row_dimensions[1].height = 36


def _truncate(s: str, max_len: int = EXCEL_MAX_CELL) -> str:
    if len(s) <= max_len:
        return s
    return s[: max_len - 40] + "\n… [truncated due to Excel cell size limit]"


def _normalize_question_type(raw: Any) -> str:
    s = str(raw or "").strip()
    if not s:
        raise SystemExit("questionsFile: Jede Frage braucht ein nicht-leeres 'type' (Subjective | Objective).")
    # Großschreibung wie in der Spezifikation (Subjective/Objective), Eingabe tolerant
    key = s[:1].upper() + s[1:].lower()
    if key == "Objective":
        return "Objective"
    if key == "Subjective":
        return "Subjective"
    raise SystemExit(
        f"questionsFile: Ungültiger type '{raw}'. Erlaubt: Subjective, Objective."
    )


def load_questions_from_file(path: Path) -> list[dict[str, Any]]:
    """Lädt Fragen aus YAML (questions: …) oder zeilenweise als Legacy-Text."""
    if path.suffix.lower() in (".yaml", ".yml"):
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise SystemExit("questionsFile (YAML): Wurzel muss ein Mapping mit Schlüssel 'questions' sein.")
        raw_list = data.get("questions")
        if not isinstance(raw_list, list):
            raise SystemExit("questionsFile (YAML): 'questions' muss eine Liste sein.")
        out: list[dict[str, Any]] = []
        for item in raw_list:
            if not isinstance(item, dict):
                raise SystemExit("questionsFile (YAML): Jedes Element unter 'questions' muss ein Objekt sein.")
            qid = item.get("id")
            qtext = item.get("question") or item.get("frage")
            if qid is None or str(qid).strip() == "":
                raise SystemExit("questionsFile (YAML): Jede Frage braucht eine 'id'.")
            if qtext is None or str(qtext).strip() == "":
                raise SystemExit(f"questionsFile (YAML): Frage {qid!r} braucht 'question'.")
            qtype = _normalize_question_type(item.get("type"))
            af = item.get("answerField")
            if af is not None and not isinstance(af, (str, int, float)):
                raise SystemExit(f"questionsFile (YAML): answerField bei {qid!r} muss Text/Zahl sein oder fehlen.")
            answer_field = "" if af is None else str(af).strip()
            out.append(
                {
                    "id": str(qid).strip(),
                    "type": qtype,
                    "question": str(qtext).strip(),
                    "answerField": answer_field or None,
                }
            )
        return out

    # Legacy: eine Frage pro Zeile
    out_legacy: list[dict[str, Any]] = []
    n = 0
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        n += 1
        out_legacy.append(
            {
                "id": str(n),
                "type": "",
                "question": line,
                "answerField": None,
            }
        )
    return out_legacy


def format_chunks_for_row(chunks: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for i, ch in enumerate(chunks, start=1):
        fn = ch.get("fileName") or ch.get("file_name") or ""
        idx = ch.get("chunkIndex", ch.get("chunk_index", ""))
        doc_id = ch.get("documentId", ch.get("document_id", ""))
        sim = ch.get("similarity", "")
        src = ch.get("source", ch.get("sourcePath", ""))
        text = str(ch.get("text", "")).strip().replace("\r\n", "\n")
        if len(text) > 2000:
            text = text[:1997] + "..."
        parts.append(
            f"--- Chunk {i} ---\n"
            f"File: {fn}\n"
            f"chunkIndex: {idx}\n"
            f"documentId: {doc_id}\n"
            f"Similarity: {sim}\n"
            f"Source: {src}\n"
            f"Text:\n{text}"
        )
    return "\n\n".join(parts) if parts else ""


def format_chunks_one_line(chunks: list[dict[str, Any]]) -> str:
    bits: list[str] = []
    for ch in chunks:
        fn = ch.get("fileName") or ch.get("file_name") or "?"
        idx = ch.get("chunkIndex", ch.get("chunk_index", "?"))
        sim = ch.get("similarity", "")
        bits.append(f"{fn}#{idx}(sim={sim})")
    return "; ".join(bits)


def merge_chat_settings(server: dict[str, Any], overlay: dict[str, Any]) -> dict[str, Any]:
    out = dict(server)
    for k, v in overlay.items():
        if v is not None:
            out[k] = v
    return out


def load_config(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise SystemExit("Konfiguration muss ein JSON-Objekt sein.")
    return data


def resolve_relative_to_config(config_file: Path, value: str) -> Path:
    """Relativ zum Verzeichnis der Konfigurationsdatei (absolute Pfade bleiben unverändert)."""
    p = Path(value)
    if p.is_absolute():
        return p
    return (config_file.resolve().parent / p).resolve()


def _active_env_display(settings: dict[str, Any]) -> tuple[str, str]:
    """Aus GET /api/settings: (activePostgresEnvironmentId, zugehöriger Anzeigename)."""
    eid = str(settings.get("activePostgresEnvironmentId") or "").strip()
    name = ""
    for env in settings.get("postgresEnvironments") or []:
        if not isinstance(env, dict):
            continue
        cid = str(env.get("id") or env.get("environmentId") or "").strip()
        if cid == eid:
            name = str(env.get("name") or "").strip()
            break
    return eid, (name if name else "—")


def print_environment_ids_for_copy(base_url: str, timeout: float) -> None:
    """Ruft GET /api/settings ab und gibt Umgebungs-IDs übersichtlich und rein zeilenweise aus (zum Kopieren)."""
    base = str(base_url or "").rstrip("/")
    with httpx.Client(base_url=base, timeout=timeout) as client:
        r = client.get("/api/settings")
        if r.status_code != 200:
            print(f"GET /api/settings fehlgeschlagen: {r.status_code} {r.text}", file=sys.stderr)
            raise SystemExit(1)
        data = r.json()
    active = str(data.get("activePostgresEnvironmentId", "") or "")
    envs = data.get("postgresEnvironments") or []
    print(f"API: {base}", flush=True)
    print("", flush=True)
    print("Umgebungen:", flush=True)
    id_lines: list[str] = []
    for env in envs:
        if not isinstance(env, dict):
            continue
        eid = str(env.get("id") or env.get("environmentId") or "").strip()
        if not eid:
            continue
        name = str(env.get("name") or "").strip()
        aktiv = " [aktiv]" if eid == active else ""
        if name:
            print(f"  • {eid} — {name}{aktiv}", flush=True)
        else:
            print(f"  • {eid}{aktiv}", flush=True)
        id_lines.append(eid)
    print("", flush=True)
    print("Nur ID (jeweils eine Zeile — in der Konsole markieren und kopieren):", flush=True)
    print("---", flush=True)
    for eid in id_lines:
        print(eid, flush=True)
    print("---", flush=True)
    print("", flush=True)
    print('JSON-Schnipsel (activePostgresEnvironmentId):', flush=True)
    if active:
        print(f'  "activePostgresEnvironmentId": "{active}"', flush=True)
    print("", flush=True)
    print("Beispiel CLI: python batch_chat.py --env <id>", flush=True)


def resolve_output_path(cfg: dict[str, Any]) -> Path:
    raw = cfg.get("outputExcel") or "batch_results.xlsx"
    p = Path(str(raw))
    if not p.is_absolute():
        p = Path.cwd() / p
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def main() -> None:
    parser = argparse.ArgumentParser(description="Batch-Chat gegen documentApi, Export nach Excel.")
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG,
        help=f"Pfad zur JSON-Konfiguration (Standard: {DEFAULT_CONFIG})",
    )
    parser.add_argument(
        "--environment",
        "--env",
        dest="environment",
        default=None,
        metavar="ID",
        help=(
            "Aktive Umgebung (activePostgresEnvironmentId), entspricht der ID in settings.json / "
            "ChatBot-Dropdown. Wird per PUT /api/settings gesetzt; überschreibt den Wert aus der Konfiguration. "
            "Verfügbare IDs: --list-environments."
        ),
    )
    parser.add_argument(
        "--list-environments",
        "--list-envs",
        action="store_true",
        help="Umgebungen von GET /api/settings auflisten (nur IDs, zum Kopieren); beendet danach ohne Batch-Lauf.",
    )
    args = parser.parse_args()
    config_path = Path(args.config)
    if not config_path.is_file():
        alt = _SCRIPT_DIR / config_path.name
        if not config_path.is_absolute() and alt.is_file():
            config_path = alt
        else:
            print(f"Konfigurationsdatei fehlt: {config_path}", file=sys.stderr)
            sys.exit(1)

    cfg = load_config(config_path)
    base = str(cfg.get("apiBaseUrl") or "http://localhost:8000").rstrip("/")
    timeout = float(cfg.get("requestTimeoutSeconds") or 300)

    if args.list_environments:
        print_environment_ids_for_copy(base, timeout)
        raise SystemExit(0)

    questions: list[dict[str, Any]] = []
    for i, q in enumerate(cfg.get("questions") or [], start=1):
        if not str(q).strip():
            continue
        questions.append(
            {
                "id": f"inline-{i}",
                "type": "",
                "question": str(q).strip(),
                "answerField": None,
            }
        )
    n_inline_json = len(questions)
    qfile = cfg.get("questionsFile")
    if not (qfile and str(qfile).strip()):
        default_yaml = config_path.parent / "question.yaml"
        if default_yaml.is_file():
            qfile = "question.yaml"

    n_from_file = 0
    qp_loaded: Path | None = None
    if qfile:
        qp = resolve_relative_to_config(config_path, str(qfile))
        if not qp.is_file():
            print(f"questionsFile nicht gefunden: {qp}", file=sys.stderr)
            sys.exit(1)
        loaded = load_questions_from_file(qp)
        questions.extend(loaded)
        n_from_file = len(loaded)
        qp_loaded = qp

    questions = [q for q in questions if str(q.get("question", "")).strip()]
    if not questions:
        print("Keine Fragen: 'questions' und/oder 'questionsFile' angeben.", file=sys.stderr)
        sys.exit(1)

    if qp_loaded is not None:
        print(
            f"Fragen geladen: {len(questions)} gesamt "
            f"({n_from_file} aus {qp_loaded.name}, {n_inline_json} inline in JSON).",
            flush=True,
        )
    else:
        print(f"Fragen geladen: {len(questions)} (nur inline aus Konfiguration).", flush=True)

    history = cfg.get("history") or []
    language = cfg.get("language")

    apply_settings = bool(cfg.get("applyChatSettings", False))
    restore_after = bool(cfg.get("restoreChatSettingsAfterRun", False))
    chat_overlay = cfg.get("chatSettings") if isinstance(cfg.get("chatSettings"), dict) else {}

    restore_app_after = bool(cfg.get("restoreAppSettingsAfterRun", False))

    app_overlay: dict[str, Any] = {}
    if isinstance(cfg.get("appSettings"), dict):
        app_overlay.update(cfg["appSettings"])
    env_top = cfg.get("activePostgresEnvironmentId") or cfg.get("active_postgres_environment_id")
    if env_top is not None and str(env_top).strip():
        app_overlay["activePostgresEnvironmentId"] = str(env_top).strip()
    cli_env = (args.environment or "").strip()
    if cli_env:
        app_overlay["activePostgresEnvironmentId"] = cli_env

    apply_app_from_flag = bool(cfg.get("applyAppSettings", False))
    apply_app_effective = apply_app_from_flag or bool(cli_env) or bool(
        env_top is not None and str(env_top).strip()
    )
    if apply_app_from_flag and not app_overlay:
        print(
            "Hinweis: applyAppSettings ist true, aber appSettings ist leer und kein activePostgresEnvironmentId — "
            "/api/settings wird nicht geändert.",
            file=sys.stderr,
            flush=True,
        )

    result_headers = _swap_results_columns_g_and_p(
        list(RESULT_HEADERS_CORE) + list(RESULT_HEADERS_AUTO) + list(RESULT_HEADERS_EVAL)
    )
    manual_prefix_col_count = len(RESULT_HEADERS_CORE) + len(RESULT_HEADERS_AUTO)
    n_auto_cols = len(RESULT_HEADERS_AUTO)

    rows: list[list[Any]] = []
    chunk_rows: list[list[Any]] = []
    chunk_headers = [
        "Question row no.",
        "Chunk position",
        "File name",
        "Chunk index",
        "Document ID",
        "Similarity",
        "Source",
        "Text excerpt",
    ]
    snapshot_before: dict[str, Any] | None = None
    snapshot_app_before: dict[str, Any] | None = None
    run_env_id = ""
    run_env_name = "—"

    with httpx.Client(base_url=base, timeout=timeout) as client:
        if apply_app_effective and app_overlay:
            ra = client.get("/api/settings")
            if ra.status_code != 200:
                print(f"GET /api/settings fehlgeschlagen: {ra.status_code} {ra.text}", file=sys.stderr)
                sys.exit(1)
            snapshot_app_before = ra.json()
            merged_app = merge_chat_settings(snapshot_app_before, app_overlay)
            rpa = client.put("/api/settings", json=merged_app)
            if rpa.status_code != 200:
                print(f"PUT /api/settings fehlgeschlagen: {rpa.status_code} {rpa.text}", file=sys.stderr)
                sys.exit(1)
            env_id = merged_app.get("activePostgresEnvironmentId", "")
            print(f"App-Einstellungen gesetzt (activePostgresEnvironmentId={env_id!r}).", flush=True)

        if apply_settings and chat_overlay:
            r0 = client.get("/api/chat/settings")
            if r0.status_code != 200:
                print(f"GET /api/chat/settings fehlgeschlagen: {r0.status_code} {r0.text}", file=sys.stderr)
                sys.exit(1)
            snapshot_before = r0.json()
            merged = merge_chat_settings(snapshot_before, chat_overlay)
            r_put = client.put("/api/chat/settings", json=merged)
            if r_put.status_code != 200:
                print(f"PUT /api/chat/settings fehlgeschlagen: {r_put.status_code} {r_put.text}", file=sys.stderr)
                sys.exit(1)

        rs_env = client.get("/api/settings")
        if rs_env.status_code == 200:
            run_env_id, run_env_name = _active_env_display(rs_env.json())
        if run_env_id:
            print(f"Batch testet gegen Umgebung: {run_env_name} (id={run_env_id})", flush=True)
        else:
            print(
                "Hinweis: Aktive Test-Umgebung nicht ermittelbar (GET /api/settings fehlgeschlagen oder leer).",
                file=sys.stderr,
                flush=True,
            )

        for i, q in enumerate(questions, start=1):
            message = str(q["question"]).strip()
            body: dict[str, Any] = {"message": message, "history": history}
            if language in ("de", "en"):
                body["language"] = language

            err_http = ""
            answer = ""
            metrics: dict[str, Any] = {}
            chunks: list[dict[str, Any]] = []

            try:
                resp = client.post("/api/chat", json=body)
                if resp.status_code != 200:
                    err_http = f"{resp.status_code}: {resp.text[:2000]}"
                else:
                    payload = resp.json()
                    answer = str(payload.get("answer", ""))
                    metrics = payload.get("metrics") or {}
                    raw_chunks = payload.get("contextChunks") or payload.get("context_chunks") or []
                    chunks = [c for c in raw_chunks if isinstance(c, dict)]
            except httpx.RequestError as exc:
                err_http = str(exc)

            af = q.get("answerField") or ""
            data_row = len(rows) + 2
            dim4_k = _dim4_kriterium_label(str(q.get("type") or ""))
            auto_placeholders: list[Any] = [""] * n_auto_cols
            rows.append(
                [
                    i,
                    q.get("id", ""),
                    q.get("type", ""),
                    message,
                    _truncate(str(af), 8000) if af else "",
                    _truncate(answer),
                    err_http,
                    metrics.get("elapsedMs", ""),
                    metrics.get("promptTokens", ""),
                    metrics.get("completionTokens", ""),
                    metrics.get("totalTokens", ""),
                    metrics.get("tokensPerSecond", ""),
                    len(chunks),
                    _truncate(format_chunks_one_line(chunks), 8000),
                    _truncate(format_chunks_for_row(chunks)),
                    *auto_placeholders,
                    "",
                    "",
                    "",
                    dim4_k,
                    "",
                    _eval_sum_formula(data_row, manual_prefix_col_count),
                ]
            )
            for pos, ch in enumerate(chunks, start=1):
                fn = ch.get("fileName") or ch.get("file_name") or ""
                idx = ch.get("chunkIndex", ch.get("chunk_index", ""))
                doc_id = ch.get("documentId", ch.get("document_id", ""))
                sim = ch.get("similarity", "")
                src = str(ch.get("source", ch.get("sourcePath", "")))
                preview = _truncate(str(ch.get("text", "")).strip().replace("\r\n", "\n"), 12000)
                chunk_rows.append([i, pos, fn, idx, doc_id, sim, src, preview])

        if restore_after and snapshot_before is not None:
            r_rest = client.put("/api/chat/settings", json=snapshot_before)
            if r_rest.status_code != 200:
                print(
                    f"Warnung: Chat-Einstellungen konnten nicht wiederhergestellt werden: "
                    f"{r_rest.status_code} {r_rest.text}",
                    file=sys.stderr,
                )

        if restore_app_after and snapshot_app_before is not None:
            r_app_rest = client.put("/api/settings", json=snapshot_app_before)
            if r_app_rest.status_code != 200:
                print(
                    f"Warnung: App-Einstellungen konnten nicht wiederhergestellt werden: "
                    f"{r_app_rest.status_code} {r_app_rest.text}",
                    file=sys.stderr,
                )

    evaluate_answers = bool(cfg.get("evaluateAnswers", False))
    if evaluate_answers:
        jc = resolve_eval_judge_config(cfg)
        print(
            f"LLM judge scoring (model {jc['llmModel']!r} @ {jc['llmBaseUrl']}) …",
            flush=True,
        )
        with httpx.Client(base_url=jc["llmBaseUrl"], timeout=jc["requestTimeoutSeconds"]) as jclient:
            for row_idx, row in enumerate(rows):
                question = str(row[3] or "")
                af = str(row[4] or "").strip()
                answer = str(row[5] or "")
                qtype = str(row[2] or "")
                detail = str(row[14] or "")
                max_cc = int(jc["maxContextChars"])
                ctx = (detail[:max_cc] if detail else None) or None
                gt = af if af else None
                judged = judge_one_http(
                    jclient,
                    api_key=jc["llmApiKey"],
                    model=jc["llmModel"],
                    temperature=float(jc["temperature"]),
                    max_tokens=int(jc["maxTokens"]),
                    question=question,
                    answer=answer,
                    context=ctx,
                    ground_truth=gt,
                    question_type=qtype,
                    max_retries=int(jc["maxRetries"]),
                    retry_sleep_sec=float(jc["retrySleepSec"]),
                )
                off = len(RESULT_HEADERS_CORE)
                row[off] = judged["answer_relevance"] if judged["answer_relevance"] is not None else ""
                row[off + 1] = judged["context_relevance"] if judged["context_relevance"] is not None else ""
                row[off + 2] = judged["groundedness"] if judged["groundedness"] is not None else ""
                # Correctness column only for Objective + reference (avoids skewed means)
                if qtype.strip() == "Objective" and af:
                    ac = judged["answer_correctness"]
                    row[off + 3] = ac if ac is not None else ""
                else:
                    row[off + 3] = ""
                row[off + 4] = judged.get("notes", "")
                row[off + 5] = _truncate(str(judged.get("error", "")), 4000)
                if (row_idx + 1) % 5 == 0:
                    print(f"  scored: {row_idx + 1}/{len(rows)}", flush=True)
        print("LLM judge done.", flush=True)

    out_path = resolve_output_path(cfg)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    meta = wb.create_sheet("Meta", 0)
    meta["A1"] = "Created (UTC)"
    meta["B1"] = datetime.now(timezone.utc).isoformat()
    meta["A2"] = "API"
    meta["B2"] = base
    meta["A3"] = "Configuration"
    meta["B3"] = str(config_path.resolve())
    meta["A4"] = "Test environment (ID)"
    meta["B4"] = run_env_id or "—"
    meta["A5"] = "Test environment (name)"
    meta["B5"] = run_env_name
    meta["A6"] = "Questions (count)"
    meta["B6"] = len(questions)
    meta["A7"] = "Scoring"
    meta["B7"] = (
        "Columns starting with “Judge:” are optional LLM-as-judge scores 0–1 (see evaluateAnswers). "
        "Columns starting with “Human score:” are manual 0–2; “Human score: total (0–8)” sums the four "
        "numeric score columns. Rubric: sheet “Rubric”; charts: sheet “Eval_Charts”."
    )
    meta["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    if apply_settings and chat_overlay:
        meta["A8"] = "Chat settings (applied)"
        meta["B8"] = json.dumps(chat_overlay, ensure_ascii=False, indent=2)
        meta["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 3):
        meta.column_dimensions[get_column_letter(col)].width = 22 if col == 1 else 80

    ws.append(result_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append(_swap_results_columns_g_and_p(list(row)))

    ws_chunks = wb.create_sheet("Chunks")
    ws_chunks.append(chunk_headers)
    for cell in ws_chunks[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for crow in chunk_rows:
        ws_chunks.append(crow)

    # Readable widths for long headers (row 1 uses wrap text).
    # Widths match post–G/P swap layout (col 7 = first judge score, col 16 = HTTP error).
    result_col_widths: dict[int, float] = {
        1: 9,
        2: 14,
        3: 16,
        4: 44,
        5: 30,
        6: 52,
        7: 30,
        8: 14,
        9: 14,
        10: 16,
        11: 14,
        12: 18,
        13: 12,
        14: 38,
        15: 48,
        16: 26,
        17: 30,
        18: 22,
        19: 32,
        20: 26,
        21: 22,
        22: 30,
        23: 28,
        24: 30,
        25: 28,
        26: 22,
        27: 22,
    }
    for col_idx in range(1, len(result_headers) + 1):
        letter = get_column_letter(col_idx)
        w = result_col_widths.get(col_idx, 14)
        ws.column_dimensions[letter].width = w
        wrap_cols = {4, 5, 6, 14, 15, 16, 20, 21}
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=col_idx)
            if col_idx in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    for jc in _JUDGE_FLOAT_DISPLAY_COLS_1BASED:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=jc)
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.number_format = _JUDGE_FLOAT_NUMBER_FORMAT

    for col_idx in range(1, len(chunk_headers) + 1):
        letter = get_column_letter(col_idx)
        chunk_widths = {1: 16, 2: 14, 3: 28, 4: 12, 5: 14, 6: 12, 7: 28, 8: 70}
        w = chunk_widths.get(col_idx, 14)
        ws_chunks.column_dimensions[letter].width = w
        if col_idx == 8:
            for r in range(2, ws_chunks.max_row + 1):
                ws_chunks.cell(row=r, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    _append_rubric_sheet(wb, insert_at=1)

    auto_eval_agg: dict[str, Any] | None = None
    if evaluate_answers and rows:
        auto_eval_agg = aggregate_auto_eval_from_rows(
            rows, core_col_count=len(RESULT_HEADERS_CORE)
        )
    if evaluate_answers and ws.max_row >= 2 and auto_eval_agg is not None:
        _append_eval_diagram_sheet(wb, aggregated=auto_eval_agg)

    eval_score_cols = _eval_score_column_indices_1based(manual_prefix_col_count)
    if ws.max_row >= 2:
        end_row = ws.max_row
        for col in eval_score_cols:
            letter = get_column_letter(col)
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=0,
                formula2=2,
                allow_blank=True,
            )
            dv.error = "Only 0, 1, or 2 (or leave blank)."
            dv.errorTitle = "Score"
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{end_row}")

    _print_auto_eval_accuracy_summary(auto_eval_agg, had_judge=evaluate_answers)
    wb.save(out_path)
    print(f"Excel written: {out_path}", flush=True)


if __name__ == "__main__":
    main()
